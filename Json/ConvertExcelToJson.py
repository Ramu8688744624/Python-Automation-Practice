from openpyxl import load_workbook
import json
try:
    wb=load_workbook("D:/Python for Automation/Excel Automation/Family.xlsx")
    sheet=wb.active
    
    data_list=[]
    
    for row in sheet.iter_rows(min_row=2,values_only=True):
        try:
            name,age,salary=row
            
            sal=int(salary)
            
            data={
                "Name":name,
                "Age":age,
                "Salary":sal
            }
            
            data_list.append(data)
        except(ValueError,TypeError):
            print("Skipping bad Row:",row)
    with open("D:/Python for Automation/Json/family.json","w") as f:
        json.dump(data_list,f,indent=4)
    print("Excel successfully converted to JSON")
        
except FileNotFoundError:
    print("File not found")
    