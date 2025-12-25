from openpyxl import load_workbook,Workbook
try:
    wb=load_workbook("D:/Python for Automation/Excel Automation/Family.xlsx")
    sheet=wb.active
    clean_data=Workbook()
    clean_sheet=clean_data.active
    clean_sheet.append(["Name","Age","Salary"])
    high_salary=Workbook()
    high_sheet=high_salary.active
    high_sheet.append(["Name","Age","Salary"])
    
    total_salary = 0
    max_salary = 0
    min_salary = 10**100
    count_of_valid_rows = 0
    
    for row in sheet.iter_rows(min_row=2,values_only=True):
        try:
            name, age, salary = row
            salary = int(salary)
            clean_sheet.append(row)
            if salary > 30000:
                high_sheet.append(row)
            
            total_salary = total_salary + salary
            count_of_valid_rows = count_of_valid_rows + 1
            
            if salary > max_salary :
                max_salary = salary
            
            if salary < min_salary:
                min_salary = salary
                
        except(ValueError,TypeError):
            print("Skipping bad row:", row)
        
    print("min_salary: ",min_salary)
    print("max_salary: ",max_salary)
    print("count_of_valid_rows: ",count_of_valid_rows)
    print("total_salary: ",total_salary)
    
    clean_data.save("D:/Python for Automation/Excel Automation/clean_data.xlsx.xlsx")
    
    high_salary.save("D:/Python for Automation/Excel Automation/high_salary.xlsx.xlsx")
except FileNotFoundError:
    print("File Not Found Exception Thrown")
	

