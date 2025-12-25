from openpyxl import load_workbook
try:
    wb=load_workbook("D:/Python for Automation/Excel Automation/Family.xlsx")


    sheet=wb.active

    print(sheet["A1"].value)
    print(sheet["B2"].value)
    print(sheet["C3"].value)
    
    for row in sheet.iter_rows(min_row=2,values_only=True):
        sal=int(row[2])
        if sal>30000:
            print(row)
except FileNotFoundError:
    print("File Not Found")
	