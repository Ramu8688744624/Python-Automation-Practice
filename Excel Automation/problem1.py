from openpyxl import load_workbook

wb = load_workbook("D:/Python for Automation/Excel Automation/Family.xlsx")
sheet=wb.active
print(sheet)
# used to read only a single cell
print(sheet["A1"].value)

for row in sheet.iter_rows(min_row=2,values_only=True):
	print(row)